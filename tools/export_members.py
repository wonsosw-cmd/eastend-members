# -*- coding: utf-8 -*-
"""회원 로우데이터를 컴퓨터(공유드라이브)로 내려받아 엑셀로 저장.

개인정보는 GitHub에 올라가지 않으며, 이 스크립트가 GAS API에서 받아
공유드라이브 '★오프라인 회원관리\회원DB' 폴더에 xlsx로 저장한다.

사용법:
    1) tools/local_config.json 생성 (git에 올라가지 않음):
       { "api_url": "https://script.google.com/macros/s/.../exec",
         "token": "관리자토큰" }
    2) py -3.14 tools/export_members.py
"""
import json
import urllib.request
import urllib.parse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUT_DIR = Path(r"G:\공유 드라이브\영업마케팅실\★. 데일리 매출 보고\★오프라인 회원관리\회원DB")

COLUMNS = [
    ("가입일시", "joinedAt", 18), ("브랜드", "brand", 10), ("매장코드", "storeId", 9),
    ("매장명", "storeName", 16), ("이름", "name", 10), ("휴대전화", "phone", 15),
    ("생년월일", "birth", 10), ("성별", "gender", 7), ("개인정보동의", "consentPrivacy", 12),
    ("마케팅동의", "consentMarketing", 11), ("최근방문", "lastVisit", 12), ("방문횟수", "visitCount", 9),
]


def main():
    cfg = json.loads((ROOT / "local_config.json").read_text(encoding="utf-8"))
    url = f"{cfg['api_url']}?action=list&token={urllib.parse.quote(cfg['token'])}"
    with urllib.request.urlopen(url, timeout=60) as r:
        data = json.loads(r.read().decode("utf-8"))
    if data.get("result") != "ok":
        raise SystemExit(f"API 오류: {data.get('message')}")
    rows = data["rows"]

    wb = Workbook()
    ws = wb.active
    ws.title = "회원"
    header_fill = PatternFill("solid", fgColor="1B2A4A")
    for c, (title, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, title)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(c)].width = width
    for r, row in enumerate(rows, 2):
        for c, (_, key, _) in enumerate(COLUMNS, 1):
            v = row.get(key, "")
            if key in ("joinedAt", "lastVisit"):
                v = str(v)[:16].replace("T", " ")
            ws.cell(r, c, v)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows) + 1, 2)}"

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / f"회원목록_{datetime.now():%Y%m%d}.xlsx"
    wb.save(out)
    print(f"저장 완료: {out}  (회원 {len(rows)}명)")


if __name__ == "__main__":
    main()
